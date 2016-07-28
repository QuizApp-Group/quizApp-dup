"""Tests for core views.
"""
import tempfile

from openpyxl import load_workbook
from factory import create_batch

from quizApp import db
from quizApp import models

from tests.auth import login_experimenter, get_participant
from tests.factories import ActivityFactory, MediaItemFactory, \
    create_experiment, DatasetFactory
from tests.helpers import json_success


def test_import_template(client, users):
    """Make sure we can get an import template.
    """
    login_experimenter(client)

    response = client.get("/import_template")
    outfile = tempfile.TemporaryFile()
    outfile.write(response.data)

    workbook = load_workbook(outfile)

    assert len(workbook.get_sheet_names())


def test_export_template(client, users):
    """Verify generation of exported data.
    """
    participant = get_participant()

    db.session.add(create_experiment(4, [participant]))
    db.session.add_all(create_batch(MediaItemFactory, 10))
    db.session.add_all(create_batch(ActivityFactory, 10))
    db.session.add_all(create_batch(DatasetFactory, 10))
    db.session.commit()

    login_experimenter(client)

    response = client.get("/export")
    outfile = tempfile.TemporaryFile()
    outfile.write(response.data)

    workbook = load_workbook(outfile)

    assert len(workbook.get_sheet_names())

    sheet_name_mapping = {"Experiments": models.Experiment.query.count(),
                          "Media items": models.MediaItem.query.count(),
                          "Activities": models.Activity.query.count(),
                          "Datasets": models.Dataset.query.count()}

    for sheet, num_objects in sheet_name_mapping.items():
        worksheet = workbook.get_sheet_by_name(sheet)
        # We should have as many rows as objects plus a header row
        assert len(worksheet.rows) == num_objects + 1


def test_import_assignments(client, users):
    login_experimenter(client)
    url = "/import"

    response = client.post(url,
                           data={"data":
                                 open("tests/data/import.xlsx")})
    assert response.status_code == 200
    assert json_success(response.data)

    assert models.Experiment.query.count() == 4
    assert models.ParticipantExperiment.query.count() == 180
    assert models.Dataset.query.count() == 6
    assert models.MediaItem.query.count() == 18
    assert models.Assignment.query.count() == 811
    assert models.Activity.query.count() == 31

    response = client.post(url)
    assert response.status_code == 200
    assert not json_success(response.data)
