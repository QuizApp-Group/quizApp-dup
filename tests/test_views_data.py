from builtins import range
import tempfile

from openpyxl import load_workbook
from factory import create_batch
import mock
import pytest

from quizApp import db
from quizApp import models

from tests.auth import login_experimenter
from tests.factories import ActivityFactory, MediaItemFactory, \
    create_experiment, DatasetFactory, ExperimentFactory
from tests.helpers import json_success

from quizApp.views import data


def test_import_template(client, users):
    """Make sure we can get an import template.
    """
    login_experimenter(client)

    response = client.get("/data/import_template")
    outfile = tempfile.TemporaryFile()
    outfile.write(response.data)

    workbook = load_workbook(outfile)

    assert len(workbook.get_sheet_names())


def test_export_template(client, users):
    """Verify generation of exported data.
    """
    db.session.add(create_experiment(4, 1))
    db.session.add_all(create_batch(MediaItemFactory, 10))
    db.session.add_all(create_batch(ActivityFactory, 10))
    db.session.add_all(create_batch(DatasetFactory, 10))
    db.session.commit()

    login_experimenter(client)

    response = client.get("/data/export")
    outfile = tempfile.TemporaryFile()
    outfile.write(response.data)

    workbook = load_workbook(outfile)

    assert len(workbook.get_sheet_names())

    sheet_name_mapping = {"Experiments": models.Experiment.query.count(),
                          "Media items": models.MediaItem.query.count(),
                          "Activities": models.Activity.query.count(),
                          "Datasets": models.Dataset.query.count()}

    for sheet, num_objects in list(sheet_name_mapping.items()):
        worksheet = workbook.get_sheet_by_name(sheet)
        # We should have as many rows as objects plus a header row
        assert len(worksheet.rows) == num_objects + 1


def test_import_assignments(client, users):
    login_experimenter(client)
    url = "/data/import"
    experiment = ExperimentFactory(id=1)
    db.session.add(experiment)

    for i in range(1, 4):
        media_item = MediaItemFactory(id=i)
        db.session.add(media_item)

    for i in range(1, 5):
        activity = ActivityFactory(id=i)
        db.session.add(activity)

    db.session.commit()

    response = client.post(url,
                           data={"data":
                                 open("tests/data/import.xlsx", "rb")})
    assert response.status_code == 200
    assert json_success(response.data)

    assert models.Experiment.query.count() == 1
    assert models.AssignmentSet.query.count() == 4
    assert models.Assignment.query.count() == 12
    assert len(models.Experiment.query.one().assignment_sets) == 4

    for pe in models.Experiment.query.one().assignment_sets:
        assert len(pe.assignments) == 3

    response = client.post(url)
    assert response.status_code == 200
    assert not json_success(response.data)

    with mock.patch("quizApp.views.data."
                    "import_data_from_workbook") as p:
        p.side_effect = AttributeError("foo")
        response = client.post(url,
                               data={"data":
                                     open("tests/data/import.xlsx", "rb")})
        assert not json_success(response.data)


def test_manage_form(client, users):
    login_experimenter(client)
    url = "/data/manage"

    response = client.get(url)

    assert response.status_code == 200


def test_header_to_field_name():
    result = data.header_to_field_name("", models.Activity)
    assert result is None

    header = "fake_tablename:foobar"

    with pytest.raises(ValueError):
        data.header_to_field_name(header, models.Activity)


def test_get_object_from_id():
    with pytest.raises(ValueError):
        data.get_object_from_id(models.Activity, 5, {})
