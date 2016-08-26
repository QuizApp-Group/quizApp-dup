"""This blueprint takes care of rendering static pages outside of the other
blueprints.
"""
import os
from collections import OrderedDict
import tempfile
import traceback

import openpyxl
from flask import Blueprint, render_template, send_file, jsonify, redirect, \
    url_for
from flask_security import roles_required, login_required, current_user

from quizApp import models, db
from quizApp.views import import_export
from quizApp.forms.core import ImportDataForm


core = Blueprint("core", __name__, url_prefix="/")


# homepage
@core.route('')
def home():
    """Display the homepage."""
    return render_template('core/index.html',
                           is_home=True)


@core.route('export')
@roles_required("experimenter")
def export_data():
    """Send the user a breakddown of datasets, activities, etc. for use in
    making assignments.
    """
    file_name = import_export.export_to_workbook()
    return send_file(file_name, as_attachment=True,
                     attachment_filename="quizapp_export.xlsx")


@core.route('import_template')
@roles_required("experimenter")
def import_template():
    """Send the user a blank excel sheet that can be filled out and used to
    populate an experiment's activity list.

    The process is essentially:

    1. Get a list of models to include

    2. From each model, get all its polymorphisms

    3. For each model, get all fields that should be included in the import
    template, including any fields from polymorphisms

    4. Create a workbook with as many sheets as models, with one row in each
    sheet, containing the name of the included fields
    """

    sheets = OrderedDict([
        ("Assignment Sets", models.AssignmentSet),
        ("Assignments", models.Assignment),
    ])

    documentation = [
        ["Do not modify the first row in every sheet!"],
        ["Simply add in your data in the rows undeneath it."],
        ["Use IDs from the export sheet to populate relationship columns."],
        [("If you want multiple objects in a relation, separate the IDs using"
          " commas.")],
    ]

    workbook = openpyxl.Workbook()
    workbook.remove_sheet(workbook.active)

    current_sheet = workbook.create_sheet()
    current_sheet.title = "Documentation"
    import_export.write_list_to_sheet(documentation, current_sheet)

    for sheet_name, model in sheets.items():
        current_sheet = workbook.create_sheet()
        current_sheet.title = sheet_name
        headers = import_export.model_to_sheet_headers(model)
        import_export.write_list_to_sheet(headers, current_sheet)

    file_handle, file_name = tempfile.mkstemp(".xlsx")
    os.close(file_handle)
    workbook.save(file_name)
    return send_file(file_name, as_attachment=True,
                     attachment_filename="import_template.xlsx")


@core.route('import', methods=["POST"])
@roles_required("experimenter")
def import_data():
    """Given an uploaded spreadsheet, import data from the spreadsheet
    into the database.
    """
    import_data_form = ImportDataForm()

    if not import_data_form.validate():
        return jsonify({"success": 0, "errors": import_data_form.errors})

    workbook = openpyxl.load_workbook(import_data_form.data.data)

    try:
        import_export.import_data_from_workbook(workbook)
    except Exception as e:
        # This isn't very nice, but we need a way to capture exceptions that
        # happen during import and show them to the user. However, we also want
        # to have the traceback for debugging purposes. So we print the
        # traceback to stdout.
        print(traceback.format_exc())
        return jsonify({"success": 0,
                        "errors": (type(e).__name__ + ": " + str(e) + "<br>"
                                   + traceback.format_exc().
                                   replace("\n", "<br>"))})

    return jsonify({"success": 1})


@core.route('manage_data', methods=["GET"])
@roles_required("experimenter")
def manage_data():
    """Show a form for uploading data and such.
    """

    import_data_form = ImportDataForm()

    return render_template("core/manage_data.html",
                           import_data_form=import_data_form)


def query_exists(query):
    """Given a query, return True if it would return any rows, and False
    otherwise.
    """
    return db.session.query(query.exists()).scalar()


@core.route("getting_started", methods=["GET"])
@roles_required("experimenter")
def getting_started():
    """Show some instructions for getting started with quizApp.
    """
    experiments_done = query_exists(models.Experiment.query)
    activities_done = query_exists(models.Activity.query)
    datasets_done = query_exists(models.Dataset.query)
    media_items_done = query_exists(models.MediaItem.query)
    assignments_done = query_exists(models.Assignment.query)

    return render_template("core/getting_started.html",
                           experiments_done=experiments_done,
                           activities_done=activities_done,
                           datasets_done=datasets_done,
                           media_items_done=media_items_done,
                           assignments_done=assignments_done)


@core.route("post_login", methods=["GET"])
@login_required
def post_login():
    """Once a user has logged in, redirect them based on their role.
    """
    if current_user.has_role("experimenter"):
        return redirect(url_for("core.getting_started"))
    else:
        return redirect(url_for("experiments.read_experiments"))
