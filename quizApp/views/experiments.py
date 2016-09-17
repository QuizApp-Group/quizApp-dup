"""Views that handle CRUD for experiments and rendering questions for
participants.
"""
from collections import defaultdict, Counter
from datetime import datetime
import json
import os
import tempfile

import openpyxl
import dateutil.parser
from flask import Blueprint, render_template, url_for, jsonify, abort, \
    request, session, send_file
from flask_security import login_required, current_user, roles_required

from quizApp import db
from quizApp.forms.experiments import CreateExperimentForm, \
    get_answer_form
from quizApp.views.common import ObjectCollectionView, ObjectView
from quizApp.models import Experiment, Assignment, \
    AssignmentSet, Participant
from quizApp.views.helpers import validate_model_id, get_first_assignment,\
    url_code_to_experiment, experiment_to_url_code
from quizApp.views.activities import render_activity
from quizApp.views.mturk import submit_assignment

experiments = Blueprint("experiments", __name__, url_prefix="/experiments")

EXPERIMENT_ROUTE = "/<int:experiment_id>"
ASSIGNMENT_SETS_ROUTE = EXPERIMENT_ROUTE + "/assignment_sets/"
ASSIGNMENT_SET_ROUTE = ASSIGNMENT_SETS_ROUTE + "<int:assignment_set_id>"
ASSIGNMENTS_ROUTE = ASSIGNMENT_SET_ROUTE + "/assignments/"
ASSIGNMENT_ROUTE = ASSIGNMENTS_ROUTE + "<int:assignment_id>"

POST_FINALIZE_HANDLERS = {
    "mturk": submit_assignment,
}


def validate_assignment_set(experiment_id, assignment_set_id):
    """Check if this experiment and assignment set exist, if this assignment
    set is part of this experiment, and if the current user owns the assignment
    set.
    """
    experiment = validate_model_id(Experiment, experiment_id)
    assignment_set = validate_model_id(AssignmentSet, assignment_set_id)

    if assignment_set.experiment != experiment:
        abort(404)

    if assignment_set.participant != current_user:
        abort(403)

    return (experiment, assignment_set)


def validate_assignment(experiment_id, assignment_set_id, assignment_id):
    """Do everything ``validate_assignment_set`` does, but also check that the
    assignment exists and that it's part of the given assignment set.
    """
    experiment, assignment_set = validate_assignment_set(experiment_id,
                                                         assignment_set_id)
    assignment = validate_model_id(Assignment, assignment_id)

    if assignment.assignment_set != assignment_set:
        abort(404)

    return (experiment, assignment_set, assignment)


class ExperimentCollectionView(ObjectCollectionView):
    """View for a collection of Experiments.
    """
    decorators = [roles_required("experimenter")]
    methods = ["GET", "POST"]
    template = "experiments/read_experiments.html"

    def create_form(self):
        return CreateExperimentForm(request.form)

    def resolve_kwargs(self, **kwargs):
        return {}

    def get_members(self):
        now = datetime.now()
        return {
            "past_experiments": Experiment.query.filter(
                Experiment.stop < now).all(),
            "present_experiments": Experiment.query.filter(
                Experiment.stop > now).filter(Experiment.start < now).all(),
            "future_experiments": Experiment.query.filter(
                Experiment.start > now),
        }

    def create_member(self, create_form):
        experiment = Experiment()
        create_form.populate_obj(experiment)
        experiment.created = datetime.now()
        experiment.save()

        return {
            "next_url": url_for("experiments.settings_experiment",
                                experiment_id=experiment.id),
        }


experiments.add_url_rule(
    "/",
    view_func=ExperimentCollectionView.as_view('experiments'))


@experiments.route("/<string:experiment_code>", methods=["GET"])
@login_required
def read_coded_experiment(experiment_code):
    """This endpoint is where participants begin an experiment. Allows them to
    start an assignment set.
    """
    experiment = url_code_to_experiment(experiment_code)

    if current_user.has_role("participant"):
        if not experiment.running:
            abort(400)
        assignment = get_first_assignment(experiment)
    else:
        assignment = None

    return render_template("experiments/read_experiment.html",
                           experiment=experiment,
                           assignment=assignment)


class ExperimentView(ObjectView):
    """View for a particular experiment.
    """
    decorators = [roles_required("experimenter")]
    methods = ["GET", "PUT", "DELETE"]
    object_key = "experiment"
    template = "experiments/read_experiment.html"

    def resolve_kwargs(self, experiment_id):
        return {"experiment": validate_model_id(Experiment, experiment_id)}

    def update_form(self, **_):
        return CreateExperimentForm(request.form)

    def collection_url(self, **_):
        return url_for("experiments.experiments")

    def get(self, experiment):
        return render_template("experiments/read_experiment.html",
                               experiment=experiment)

experiments.add_url_rule(
    EXPERIMENT_ROUTE,
    view_func=ExperimentView.as_view('experiment'))


@experiments.route(ASSIGNMENT_ROUTE, methods=["GET"])
@roles_required("participant")
def read_assignment(experiment_id, assignment_set_id, assignment_id):
    """Given an assignment ID, retrieve it from the database and display it to
    the user.
    """
    experiment, assignment_set, assignment = validate_assignment(
        experiment_id,
        assignment_set_id,
        assignment_id)

    if not experiment.running:
        abort(400)

    if experiment.disable_previous and assignment_set.progress > \
            assignment_set.assignments.index(assignment) and \
            not assignment_set.complete:
        abort(400)

    activity = assignment.activity

    read_function_mapping = {
        "question_mc_singleselect": read_question,
        "question_mc_multiselect": read_question,
        "question_freeanswer": read_question,
        "question_mc_singleselect_scale": read_question,
        "question_integer": read_question,
        "scorecard": read_scorecard,
    }

    return read_function_mapping[activity.type](experiment, assignment)


def read_scorecard(experiment, assignment):
    """Read an assignment that is a scorecard.
    """
    scorecard = assignment.activity
    scorecard_form = get_answer_form(scorecard)
    scorecard_form.populate_from_assignment(assignment)

    assignment_set = assignment.assignment_set
    this_index = assignment_set.assignments.index(assignment)

    next_url = get_next_assignment_url(assignment_set, this_index)

    # Get the previous assignment, if any
    previous_assignment = None
    if this_index - 1 > -1 and not experiment.disable_previous:
        previous_assignment = assignment_set.assignments[this_index - 1]

    scorecard = assignment.activity
    cumulative_score = assignment.assignment_set.score
    rendered_scorecard = render_activity(scorecard, assignment_set.complete,
                                         assignment_set, assignment,
                                         this_index)

    template_kwargs = {
        "exp": experiment,
        "assignment": assignment,
        "next_url": next_url,
        "cumulative_score": cumulative_score,
        "experiment_complete": assignment_set.complete,
        "previous_assignment": previous_assignment,
        "rendered_scorecard": rendered_scorecard,
        "scorecard_form": scorecard_form,
    }

    return render_template("experiments/read_scorecard.html",
                           **template_kwargs)


def read_question(experiment, assignment):
    """Common code for reading questions.
    """
    question = assignment.activity
    question_form = get_answer_form(question)
    question_form.populate_from_assignment(assignment)

    assignment_set = assignment.assignment_set
    this_index = assignment_set.assignments.index(assignment)

    if assignment.result:
        question_form.populate_from_result(assignment.result)

    if not assignment_set.complete:
        next_url = None
    else:
        # If the participant is done, have a link right to the next question
        next_url = get_next_assignment_url(assignment_set, this_index)

    # Get the previous assignment, if any
    previous_assignment = None
    if this_index - 1 > -1 and not experiment.disable_previous:
        previous_assignment = assignment_set.assignments[this_index - 1]

    cumulative_score = assignment.assignment_set.score
    rendered_question = render_activity(question, assignment_set.complete,
                                        assignment, assignment_set.complete)

    template_kwargs = {
        "exp": experiment,
        "assignment": assignment,
        "question_form": question_form,
        "next_url": next_url,
        "cumulative_score": cumulative_score,
        "experiment_complete": assignment_set.complete,
        "previous_assignment": previous_assignment,
        "rendered_question": rendered_question,
    }

    # This mapping is for further processing of certain question types, if
    # necessary
    read_question_function_mapping = {
        "question_mc_singleselect": read_mc_question,
        "question_mc_multiselect": read_mc_question,
        "question_mc_singleselect_scale": read_mc_question,
    }

    try:
        read_question_function_mapping[question.type](experiment, assignment)
    except KeyError:
        pass

    return render_template("experiments/read_question.html",
                           **template_kwargs)


def read_mc_question(_, assignment):
    """Read a multiple choice question, making sure to save the choice order.
    """
    if not assignment.assignment_set.complete:
        # If the participant is not done, then save the choice order
        choice_order = [c.id for c in assignment.activity.choices]
        assignment.choice_order = json.dumps(choice_order)
        assignment.save()


@experiments.route(ASSIGNMENT_ROUTE, methods=["PATCH"])
def update_assignment(experiment_id, assignment_set_id, assignment_id):
    """Record a user's answer to this assignment
    """
    experiment, assignment_set, assignment = validate_assignment(
        experiment_id,
        assignment_set_id,
        assignment_id)

    if assignment_set.complete:
        abort(400)

    if not experiment.running:
        abort(400)

    if experiment.disable_previous and assignment_set.progress > \
            assignment_set.assignments.index(assignment):
        abort(400)

    activity_form = get_answer_form(assignment.activity, request.form)
    activity_form.populate_from_activity(assignment.activity)

    if not activity_form.validate():
        return jsonify({"success": 0, "errors": activity_form.errors})

    activity_form.populate_assignment(assignment)

    this_index = assignment_set.assignments.index(assignment)
    next_url = get_next_assignment_url(assignment_set, this_index)

    if this_index == assignment_set.progress:
        assignment_set.progress += 1

    # Record time to solve
    if activity_form.render_time.data and activity_form.submit_time.data:
        render_datetime = dateutil.parser.parse(activity_form.render_time.data)
        submit_datetime = dateutil.parser.parse(activity_form.submit_time.data)
        time_to_submit = submit_datetime - render_datetime
        assignment.time_to_submit = time_to_submit

    db.session.commit()

    if assignment.activity.scorecard_settings.display_scorecard:
        return jsonify({
            "success": 1,
            "scorecard": render_template(
                "experiments/interim_scorecard.html",
                scorecard_settings=assignment.activity.scorecard_settings,
                assignment=assignment,
                next_url=next_url)
        })

    return jsonify({"success": 1, "next_url": next_url})


def get_next_assignment_url(assignment_set, current_index):
    """Given an experiment, a assignment_set, and the current index,
    find the url of the next assignment in the sequence.
    """
    experiment_id = assignment_set.experiment.id
    try:
        # If there is a next assignment, return its url
        next_url = url_for(
            "experiments.read_assignment",
            experiment_id=experiment_id,
            assignment_set_id=assignment_set.id,
            assignment_id=assignment_set.assignments[current_index + 1].id)
    except IndexError:
        next_url = None

    if not next_url:
        # We've reached the end of the experiment
        if not assignment_set.complete:
            # The experiment needs to be submitted
            next_url = url_for(
                "experiments.confirm_done_assignment_set",
                assignment_set_id=assignment_set.id,
                experiment_code=experiment_to_url_code(
                    assignment_set.experiment),
                experiment_id=experiment_id)
        else:
            # Experiment has already been submitted
            next_url = url_for("experiments.experiment",
                               experiment_id=experiment_id)

    return next_url


@experiments.route(EXPERIMENT_ROUTE + '/settings', methods=["GET"])
@roles_required("experimenter")
def settings_experiment(experiment_id):
    """Give information on an experiment and its activities.
    """
    experiment = validate_model_id(Experiment, experiment_id)

    update_experiment_form = CreateExperimentForm(obj=experiment)
    experiment_code = experiment_to_url_code(experiment)
    coded_url = url_for("experiments.read_coded_experiment",
                        experiment_code=experiment_code, _external=True)

    return render_template("experiments/settings_experiment.html",
                           experiment=experiment,
                           coded_url=coded_url,
                           update_experiment_form=update_experiment_form)


def get_question_stats(assignment, question_stats):
    """Given an assignment of a question and a stats array, return statistics
    about this question in the array.
    """
    question = assignment.activity
    if question.id not in question_stats:
        question_stats[question.id] = {
            "num_responses": 0,
            "num_correct": 0,
            "question_text": question.question,
        }

    if assignment.result:
        question_stats[question.id]["num_responses"] += 1


@experiments.route(EXPERIMENT_ROUTE + "/results", methods=["GET"])
@roles_required("experimenter")
def results_experiment(experiment_id):
    """Render some results.
    """
    experiment = validate_model_id(Experiment, experiment_id)

    num_participants = Participant.query.count()
    num_finished = AssignmentSet.query.\
        filter_by(experiment_id=experiment.id).\
        filter_by(complete=True).count()

    percent_finished = num_finished / float(num_participants)

    # {"question_id": {"question": "question_text", "num_responses":
    #   num_responses, "num_correct": num_correct], ...}
    question_stats = defaultdict(dict)
    assignments = Assignment.query.join(AssignmentSet).\
        filter(AssignmentSet.experiment == experiment).all()

    for assignment in assignments:
        activity = assignment.activity

        if "question" in activity.type:
            get_question_stats(assignment, question_stats)

    return render_template("experiments/results_experiment.html",
                           experiment=experiment,
                           num_participants=num_participants,
                           num_finished=num_finished,
                           percent_finished=percent_finished,
                           question_stats=question_stats)


@experiments.route(EXPERIMENT_ROUTE + "/results/export", methods=["GET"])
@roles_required("experimenter")
def export_results_experiment(experiment_id):
    """Get a spreadsheet breaking down how participants did in this experiment.
    """
    # Grab all assignments in the experiment
    # For every assignment
    #  If the activity is not present in the headers
    #   Add it to the headers
    #  If the user is not present in the rows
    #   Add them to the rows
    #  Determine the activity's position in the headers
    #  Determine the user's position in the rows
    #  The above two give us coordinates, fill them out with user's answer,
    #  points, etc.
    experiment = validate_model_id(Experiment, experiment_id)

    workbook = get_results_workbook(experiment)

    file_name = tempfile.mkstemp(".xlsx")
    os.close(file_name[0])
    workbook.save(file_name[1])
    return send_file(
        file_name[1], as_attachment=True,
        attachment_filename="experiment_{}_report.xlsx".format(experiment.id))


def get_activity_column_index(activity, activity_column_mapping,
                              activity_counter, headers):
    """Find the column index for this occurrence of the given activity. This
    will update headers, counter, and mapping if necessary.
    """
    activity_occurrence = activity_counter[activity.id]
    activity_counter[activity.id] += 1
    try:
        return activity_column_mapping[activity.id][activity_occurrence]
    except KeyError:
        activity_column_mapping[activity.id] = [len(headers) + 1]
        headers.append("{}: {}".format(activity.id, activity))
        headers.append("Correct?")
        headers.append("Points")
    except IndexError:
        activity_column_mapping[activity.id].append(len(headers) + 1)
        headers.append("{}: {}".format(activity.id, activity))
        headers.append("Correct?")
        headers.append("Points")
    return activity_column_mapping[activity.id][activity_occurrence]


def get_results_workbook(experiment):
    """Analyze the assignment sets in the experiment and return an excel
    workbook.
    """
    assignment_sets = experiment.assignment_sets
    headers = ["User email", "User ID"]
    activity_column_mapping = {}
    next_participant_row = 2
    participant_row_mapping = {}

    # The same activity can appear multiple times in an assignment set. To
    # display them properly, we keep a list of their ocurrences in
    # activity_column_mapping, like so:
    # {1: [3, 7, 10], ...} means activity 1 occurs in columns 3, 7, and 10
    # When populating a row, we will use the earliest occurrence of the
    # activity possible.

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Experiment {} - Report".format(experiment.id)

    for assignment_set in assignment_sets:
        activity_counter = Counter()
        participant = assignment_set.participant

        # Nobody has done this set
        if not participant:
            continue

        # Encountered a new participant
        if participant.id not in participant_row_mapping:
            participant_row_mapping[participant.id] = next_participant_row
            populate_row_segment(sheet,
                                 next_participant_row,
                                 1,
                                 [participant.email, participant.id])
            next_participant_row += 1

        for assignment in assignment_set.assignments:
            activity = assignment.activity
            activity_column_index = get_activity_column_index(
                activity,
                activity_column_mapping,
                activity_counter,
                headers)

            if not assignment.result:
                row = ["_BLANK_"] * 3
            else:
                row = ["{}:{}".format(assignment.id, assignment.result),
                       assignment.correct,
                       assignment.score]

            populate_row_segment(
                sheet,
                participant_row_mapping[participant.id],
                activity_column_index,
                row
            )

    populate_row_segment(sheet, 1, 1, headers)

    # Specify experiment ID
    sheet.cell(row=1, column=len(headers) + 1).value = "Experiment ID"
    sheet.cell(row=2, column=len(headers) + 1).value = experiment.id

    return workbook


def populate_row_segment(sheet, row_index, initial_col, row):
    """Populate the segment of row # ``row_index`` in ``sheet`` that starts at
    ``initial_col`` and contains the items in ``row``.
    """
    for col_offset, cell in enumerate(row):
        sheet.cell(row=row_index, column=initial_col + col_offset).value = cell


@experiments.route(ASSIGNMENT_SET_ROUTE + "/confirm_done", methods=["GET"])
@roles_required("participant")
def confirm_done_assignment_set(experiment_id, assignment_set_id):
    """Show the user a page before finalizing their quiz answers.
    """
    experiment, assignment_set = validate_assignment_set(experiment_id,
                                                         assignment_set_id)

    return render_template("experiments/confirm_done_assignment_set.html",
                           assignment_set=assignment_set,
                           experiment=experiment)


@experiments.route(ASSIGNMENT_SET_ROUTE + "/finalize", methods=["PATCH"])
@roles_required("participant")
def finalize_assignment_set(experiment_id, assignment_set_id):
    """Finalize the user's answers for this experiment. They will no longer be
    able to edit them, but may view them.
    """
    experiment, assignment_set = validate_assignment_set(experiment_id,
                                                         assignment_set_id)

    if assignment_set.complete:
        abort(400)

    assignment_set.complete = True

    db.session.commit()

    return jsonify({"success": 1,
                    "next_url": url_for('experiments.done_assignment_set',
                                        assignment_set_id=assignment_set.id,
                                        experiment_id=experiment.id)})


@experiments.route(ASSIGNMENT_SET_ROUTE + "/done", methods=["GET"])
@roles_required("participant")
def done_assignment_set(experiment_id, assignment_set_id):
    """Show the user a screen indicating that they are finished.
    """
    experiment, assignment_set = validate_assignment_set(experiment_id,
                                                         assignment_set_id)

    # Handle any post finalize actions, e.g. providing a button to submit a HIT
    post_finalize = session.pop("experiment_post_finalize_handler", None)
    addendum = None
    if post_finalize:
        handler = POST_FINALIZE_HANDLERS[post_finalize]
        addendum = handler()

    return render_template("experiments/done_assignment_set.html",
                           addendum=addendum,
                           assignment_set=assignment_set,
                           scorecard_settings=experiment.scorecard_settings)
