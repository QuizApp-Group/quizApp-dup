#!/usr/bin/env python
"""Make a spreadsheet usable for importing.
"""

import openpyxl

DEST_FILE = "demo.xlsx"

# This list represents the data to be written to the sheet
# The format of this list is:
# [assignment_set1, assignment_set2, ...]

# Each assignment set in the list is represented by a tuple. The format of
# each assignment set tuple is:
# (experiment_id, [assignment1, assignment2, ...])

# Each assignment in the list is also a tuple. The format of each assignment
# tuple is:
# ([media_item1, media_item2, ...], activity1)
DATA = [
    (7, [([1], 1),
          ([2], 2),
          ([3], 3),
          ([1], 4),
          ([], 59),
          ([1], 5),
          ([2], 1),
          ([3], 2),
          ([1], 3),
          ([], 59),
          ],
     ),
]


def write_headers(prefix, headers, sheet):
    """Write out the given headers to the given sheet, using the given prefix
    for each header.
    """
    contents = ["{}:{}".format(prefix, header) for header in headers]
    write_row(1, contents, sheet)


def write_row(index, contents, sheet):
    """Given a list ``contents``, write out every element to the given
    ``sheet`` in its own cell on row ``index``.
    """
    for col, cell in enumerate(contents, 1):
        sheet.cell(row=index, column=col).value = cell


def write_data_to_workbook(data, workbook):
    """Given data, in the format described above, and a workbook, write
    ``data`` to the ``workbook``.
    """
    # Initialize our two sheets
    sheets = ["Assignment Sets", "Assignments"]

    for sheet_title in sheets:
        current_sheet = workbook.create_sheet()
        current_sheet.title = sheet_title

    assignment_set_sheet = workbook.get_sheet_by_name("Assignment Sets")
    write_headers("assignment_set",
                  ["experiment", "id", "assignments"],
                  assignment_set_sheet)

    assignment_sheet = workbook.get_sheet_by_name("Assignments")
    write_headers("assignment",
                  ["media_items", "activity", "id"],
                  assignment_sheet)

    next_assignment_id = 1
    for assignment_set_id, assignment_set in enumerate(data, 1):
        assignment_ids = []

        for assignment in assignment_set[1]:
            row = [",".join([str(a) for a in assignment[0]]),
                   assignment[1],
                   next_assignment_id,
                   ]
            assignment_ids.append(str(next_assignment_id))
            next_assignment_id += 1
            write_row(next_assignment_id, row, assignment_sheet)

        row = [assignment_set[0],
               assignment_set_id,
               ",".join(assignment_ids),
               ]
        write_row(assignment_set_id + 1, row, assignment_set_sheet)


def main():
    """Write out ``DATA`` to a clean workbook.
    """
    workbook = openpyxl.Workbook()

    # workbooks are instantiated with one default sheet - we remove it so that
    # our loop later is a bit cleaner
    workbook.remove_sheet(workbook.active)

    write_data_to_workbook(DATA, workbook)

    # Write the workbook out to disk
    workbook.save(DEST_FILE)


if __name__ == "__main__":
    main()
